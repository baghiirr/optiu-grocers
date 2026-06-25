from __future__ import annotations

import os
from datetime import datetime, timezone

from openpyxl import Workbook

SHEET_ORDER = [
    "SKUs",
    "Suppliers",
    "Vendor terms",
    "Inventory snapshots",
    "Invoices",
    "Invoice lines",
    "Stores",
    "Employees",
    "Labor rules",
]

SHEET_COLUMNS: dict[str, list[str]] = {
    "SKUs": [
        "sku_id", "sku_name", "upc", "brand", "category", "subcategory",
        "pack_size", "case_qty", "is_age_restricted", "is_lottery", "is_food",
    ],
    "Suppliers": ["vendor_id", "vendor_name", "delivery_days", "edi_id"],
    "Vendor terms": ["sku_id", "vendor_id", "case_cost", "case_qty", "min_order_units", "lead_time_days"],
    "Inventory snapshots": ["site_id", "sku_id", "snapshot_at", "on_hand_units", "shelf_price"],
    "Invoices": [
        "invoice_id", "site_id", "vendor_id", "invoice_date", "total_amount",
        "source_channel", "confidence_score", "received_at",
    ],
    "Invoice lines": [
        "invoice_id", "line_num", "sku_id", "description", "units",
        "case_cost", "line_total", "confidence_score",
    ],
    "Stores": [
        "site_id", "site_name", "street_address", "phone", "city", "state", "zip",
        "format", "pump_count", "store_sqft", "has_carwash", "has_food_program",
        "has_lottery", "has_alcohol", "open_24h", "open_date", "is_active",
        "timezone", "latitude", "longitude",
    ],
    "Employees": ["employee_id", "full_name", "site_id", "role", "hourly_rate", "hired_on", "is_active"],
    "Labor rules": [
        "rule_id", "state", "min_shift_hours", "max_weekly_hours",
        "overtime_after", "overtime_rate_mul", "break_minutes",
    ],
}


def _cell_value(value):
    if isinstance(value, bool):
        return "true" if value else "false"
    return value


def _write_catalog_export_sheet(wb: Workbook, rows_by_sheet: dict[str, list[dict]]) -> None:
    ws = wb.create_sheet("Catalog export")
    generated_at = datetime.now(tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    ws.append(["OptiU — Catalog data export"])
    ws.append([])
    ws.append(["Generated", generated_at])
    ws.append(["Source", "Clover POS sync"])
    ws.append([])
    ws.append(["Sheet", "Rows"])
    for sheet_name in SHEET_ORDER:
        ws.append([sheet_name, len(rows_by_sheet.get(sheet_name, []))])


def build_workbook(rows_by_sheet: dict[str, list[dict]], output_path: str | os.PathLike) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _write_catalog_export_sheet(wb, rows_by_sheet)

    for sheet_name in SHEET_ORDER:
        columns = SHEET_COLUMNS[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        for row in rows_by_sheet.get(sheet_name, []):
            ws.append([_cell_value(row.get(col)) for col in columns])

    wb.save(output_path)
