from __future__ import annotations

from openpyxl import load_workbook

from opti_connector.workbook import SHEET_COLUMNS, SHEET_ORDER, build_workbook


def test_workbook_has_all_required_sheets_in_order(tmp_path):
    output = tmp_path / "wb.xlsx"
    build_workbook({}, output)

    wb = load_workbook(output)
    assert wb.sheetnames == SHEET_ORDER


def test_header_and_description_rows(tmp_path):
    output = tmp_path / "wb.xlsx"
    build_workbook({}, output)

    wb = load_workbook(output)
    ws = wb["INP_STORE"]
    header = [cell.value for cell in ws[1]]
    description = [cell.value for cell in ws[2]]
    assert header == SHEET_COLUMNS["INP_STORE"]
    assert description != header
    assert any(d for d in description)


def test_data_starts_on_row_three_and_matches_input(tmp_path):
    output = tmp_path / "wb.xlsx"
    rows = {
        "INP_STORE": [
            {"store_id": "ST1", "store_name": "Test Store", "currency": "USD"},
        ]
    }
    build_workbook(rows, output)

    wb = load_workbook(output)
    ws = wb["INP_STORE"]
    assert ws.max_row == 3
    data_row = [cell.value for cell in ws[3]]
    columns = SHEET_COLUMNS["INP_STORE"]
    expected = [rows["INP_STORE"][0].get(col) for col in columns]
    assert data_row == expected


def test_boolean_written_as_lowercase_string(tmp_path):
    output = tmp_path / "wb.xlsx"
    rows = {
        "INP_SALES_HISTORY": [
            {
                "product_id": "P1",
                "store_id": "S1",
                "sale_date": "2026-06-01",
                "units_sold": 1,
                "revenue": 1.0,
                "avg_price": 1.0,
                "on_markdown": True,
            }
        ]
    }
    build_workbook(rows, output)

    wb = load_workbook(output)
    ws = wb["INP_SALES_HISTORY"]
    columns = SHEET_COLUMNS["INP_SALES_HISTORY"]
    on_markdown_idx = columns.index("on_markdown")
    value = ws[3][on_markdown_idx].value
    assert value == "true"


def test_empty_sheet_still_present_with_headers_only(tmp_path):
    output = tmp_path / "wb.xlsx"
    build_workbook({"INP_PRODUCT": []}, output)

    wb = load_workbook(output)
    ws = wb["INP_PRODUCT"]
    assert ws.max_row == 2
